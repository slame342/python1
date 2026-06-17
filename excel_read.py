# устанавливаем библиотеку для работы с эксель
# pip install openpyxl
# импортируем библиотеку для работы с эксель
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

city_list = "temp/sample_example.xlsx"

# загружаем в память уже существующий файл на диске
workbook = openpyxl.load_workbook(city_list)


# берет активную страницу из рабочей книги
worksheet = workbook.active

# берем последнюю строку в эксель файле
max_row = worksheet.max_row
print(max_row)
print(type(max_row))
# берем последнюю колонку в эксель файле
max_column = worksheet.max_column
print(max_column)
print(type(max_column))

index = 0
for i in range(1, max_row):

    # получение значения с выбранной ячейки, где row - это строка, column - это колонка
    # value = worksheet.cell(row=i, column=max_column).value

    # получение значения с выбранной ячейки,где  в квадратных скобках координаты ячейки
    value = worksheet[f"A{i}"].value
    # cell_object = worksheet[f"A{i}"]
    # value = cell_object.value
    # if value is not None:
    #     pass
    # if len(str(value)) >= 1:
    if value:
        print(value)
        print(type(value))
        index += 1
print(index)

# for i in range(1, max_row + 1):
#     for j in range(1, max_column + 1):
#
#
# # for num in var_range:
# #     for city in city_list:
# #         row = num
# #         # 1-> A, 3->C, 26->Z
# #         col = get_column_letter(city_list.index(city)+1)
# #         # col = "A"
# #         worksheet[f"{col}{row}"] = str(city)
#
# workbook.save('sample_example.xlsx')

