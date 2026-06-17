# устанавливаем библиотеку для работы с эксель
# pip install openpyxl
# импортируем библиотеку для работы с эксель
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# создание объекта из библиотеки openpyxl
workbook = Workbook()

# берет активную страницу из рабочей книги
worksheet = workbook.active

city_list = ["Almaty", "Astana", "Taraz", "Shymkent" ]
var_range = range(1, 1000, 1)
# for n in var_range:
#     print(n)
print(var_range)

# функция range возвращает массив чисел
# x = range(6)
# for n in x:
#     # print(n)
#     pass


# x = range(3, 200, 2)
# for n in x:
#     print(n)

for num in var_range:
    for city in city_list:
        row = num
        # 1-> A, 3->C, 26->Z
        col = get_column_letter(city_list.index(city)+1)
        # col = "A"
        worksheet[f"{col}{row}"] = str(city)

# "кривой способ решения"
# for j in "ABCD":
#     row = "1"
#     col = j
#     worksheet[f'{col}{row}'] = str(city_list["ABCD".index(j)])

# записываем значение в выбранную(A1) ячейку
# worksheet['A1'] = 42

# worksheet.append([1, 2, 3,])
#
# worksheet['A2'] = datetime.datetime.now()

# сохраняем рабочую книгу в excel-файл(xlsx/xls
workbook.save('sample.xlsx')

